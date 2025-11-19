"""
Excel export functionality for real estate investment model.
"""

import pandas as pd
from typing import Dict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import ModelParameters


def create_parameters_sheet(params: ModelParameters) -> pd.DataFrame:
    """Create parameters summary as DataFrame for Excel"""

    data = []

    # Header
    data.append(['APARTMENT & TRANSACTION PARAMETERS', '', ''])
    data.append(['Parameter', 'Value', 'Unit'])

    # Apartment parameters
    data.append(['Apartment Cost (USD)', params.apartment_cost_usd, 'USD'])
    data.append(['Exchange Rate (USD/UAH)', params.fx_today, 'UAH per USD'])
    data.append(['Apartment Cost (UAH)', params.apartment_cost_uah, 'UAH'])
    data.append(['Down Payment (USD)', params.downpayment_usd, 'USD'])
    data.append(['Extra Purchase Costs (USD)', params.extra_purchase_costs_usd, 'USD'])
    data.append(['Total Own Cash (USD)', params.own_cash_total_usd, 'USD'])
    data.append(['Total Own Cash (UAH)', params.own_cash_total_uah, 'UAH'])

    data.append(['', '', ''])
    data.append(['LOAN PARAMETERS', '', ''])
    data.append(['Loan Amount (UAH)', params.loan_amount_uah, 'UAH'])
    data.append(['Loan Amount (USD equivalent)', params.loan_amount_uah / params.fx_today, 'USD'])
    data.append(['Loan Term (Years)', params.loan_term_years, 'years'])
    data.append(['Loan Term (Months)', params.loan_term_months, 'months'])
    data.append(['Annual Interest Rate', params.interest_annual, 'fraction'])
    data.append(['Annual Interest Rate (%)', params.interest_annual * 100, '%'])
    data.append(['Monthly Interest Rate', params.interest_monthly, 'fraction'])
    data.append(['Annual Insurance Rate', params.insurance_annual, 'fraction'])
    data.append(['Annual Insurance Rate (%)', params.insurance_annual * 100, '%'])
    data.append(['Annual Maintenance Rate', params.maintenance_annual, 'fraction'])
    data.append(['Annual Maintenance Rate (%)', params.maintenance_annual * 100, '%'])

    data.append(['', '', ''])
    data.append(['Monthly Principal Payment', params.principal_monthly, 'UAH'])
    data.append(['Monthly Insurance Payment', params.insurance_monthly_uah, 'UAH'])
    data.append(['Monthly Maintenance Cost', params.maintenance_monthly_uah, 'UAH'])

    data.append(['', '', ''])
    data.append(['RENTAL PARAMETERS', '', ''])
    data.append(['Initial Rent (UAH)', params.rent_initial_uah, 'UAH'])
    data.append(['Initial Rent (USD)', params.rent_initial_uah / params.fx_today, 'USD'])

    data.append(['', '', ''])
    data.append(['DISCOUNT RATE', '', ''])
    data.append(['USD Discount Rate (Annual)', params.usd_discount_annual, 'fraction'])
    data.append(['USD Discount Rate (Annual %)', params.usd_discount_annual * 100, '%'])
    data.append(['USD Discount Rate (Monthly)', params.usd_discount_monthly, 'fraction'])

    data.append(['', '', ''])
    data.append(['SCENARIOS', '', ''])
    for scenario_name, scenario in params.scenarios.items():
        data.append([f'--- {scenario_name.upper()} SCENARIO ---', '', ''])
        data.append(['Rent Growth (Annual %)', scenario.rent_growth_annual * 100, '%'])
        data.append(['UAH Inflation (Annual %)', scenario.inflation_uah_annual * 100, '%'])
        data.append(['Apartment Price Growth USD (Annual %)', scenario.price_growth_annual_usd * 100, '%'])
        data.append(['', '', ''])

    return pd.DataFrame(data, columns=['Parameter', 'Value', 'Unit'])


def create_summary_sheet(all_metrics: Dict[str, Dict[str, float]]) -> pd.DataFrame:
    """Create summary sheet with metrics for all scenarios"""

    rows = []

    # Get scenario names (use actual scenario names from metrics)
    scenario_names = list(all_metrics.keys())

    # Ensure we have exactly 3 scenarios for the table
    while len(scenario_names) < 3:
        scenario_names.append('')
    scenario_names = scenario_names[:3]

    # Header
    header_row = ['INVESTMENT METRICS SUMMARY'] + [''] * len(scenario_names)
    rows.append(header_row)
    rows.append(['Metric'] + [s.capitalize() for s in scenario_names])

    # Get metrics for each scenario
    scenarios_data = [all_metrics.get(name, {}) for name in scenario_names]

    # Initial Investment
    rows.append(['', '', '', ''])
    rows.append(['INITIAL INVESTMENT', '', '', ''])
    rows.append(['Own Cash (USD)'] + [s.get('Initial_Investment_USD', '') for s in scenarios_data])

    # Rent
    rows.append(['', '', '', ''])
    rows.append(['RENTAL INCOME', '', '', ''])
    rows.append(['Total Rent Collected (Nominal USD)'] + [s.get('Total_Rent_Collected_USD_nominal', '') for s in scenarios_data])
    rows.append(['Total Rent Collected (Real USD)'] + [s.get('Total_Rent_Collected_USD_real', '') for s in scenarios_data])

    # Expenses
    rows.append(['', '', '', ''])
    rows.append(['EXPENSES', '', '', ''])
    rows.append(['Total Mortgage Paid (Nominal USD)'] + [s.get('Total_Mortgage_Paid_USD_nominal', '') for s in scenarios_data])
    rows.append(['Total Mortgage Paid (Real USD)'] + [s.get('Total_Mortgage_Paid_USD_real', '') for s in scenarios_data])
    rows.append(['Total Maintenance (Real USD)'] + [s.get('Total_Maintenance_USD_real', '') for s in scenarios_data])

    # NPV
    rows.append(['', '', '', ''])
    rows.append(['NET PRESENT VALUE', '', '', ''])
    rows.append(['NPV without Sale (Real USD)'] + [s.get('NPV_Real_USD_no_sale', '') for s in scenarios_data])
    rows.append(['NPV with Sale (Real USD)'] + [s.get('NPV_Real_USD_with_sale', '') for s in scenarios_data])

    # Terminal value
    rows.append(['', '', '', ''])
    rows.append(['TERMINAL VALUE (SALE)', '', '', ''])
    rows.append(['Sale Price (Nominal USD)'] + [s.get('Terminal_Price_USD_nominal', '') for s in scenarios_data])
    rows.append(['Sale Price (Real USD)'] + [s.get('Terminal_Price_USD_real', '') for s in scenarios_data])

    # Returns
    rows.append(['', '', '', ''])
    rows.append(['RETURNS', '', '', ''])
    rows.append(['IRR Annual (USD) %'] + [
        s.get('IRR_annual_USD_with_sale', '') if s.get('IRR_annual_USD_with_sale') is not None else 'N/A'
        for s in scenarios_data
    ])
    rows.append(['ROI (Real USD with Sale) %'] + [
        s.get('ROI_Real_USD_with_sale', '') if s.get('ROI_Real_USD_with_sale') is not None else 'N/A'
        for s in scenarios_data
    ])

    # Column names based on actual scenario names
    columns = ['Metric'] + [s.capitalize() for s in scenario_names]
    return pd.DataFrame(rows, columns=columns)


def export_to_excel(
    params: ModelParameters,
    credit_df: pd.DataFrame,
    rent_schedules: Dict[str, pd.DataFrame],
    cashflows: Dict[str, pd.DataFrame],
    all_metrics: Dict[str, Dict[str, float]],
    output_path: str
):
    """
    Export all data to Excel with multiple sheets.

    Sheets:
    - Parameters: All input parameters
    - Summary: Key metrics for all scenarios
    - Credit_Schedule: Mortgage payment schedule
    - Rent_Schedule_XXX: Rent schedules for each scenario
    - Cashflow_XXX: Cashflow analysis for each scenario
    """

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # Parameters sheet
        params_df = create_parameters_sheet(params)
        params_df.to_excel(writer, sheet_name='Parameters', index=False)

        # Summary sheet
        summary_df = create_summary_sheet(all_metrics)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Credit schedule
        credit_df.to_excel(writer, sheet_name='Credit_Schedule', index=False)

        # Rent schedules for each scenario
        for scenario_name, rent_df in rent_schedules.items():
            sheet_name = f'Rent_{scenario_name.capitalize()}'
            rent_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Cashflow for each scenario
        for scenario_name, cashflow_df in cashflows.items():
            sheet_name = f'Cashflow_{scenario_name.capitalize()}'
            cashflow_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply formatting
    _format_excel(output_path)

    print(f"Excel file exported successfully: {output_path}")


def _format_excel(filepath: str):
    """Apply formatting to Excel file"""

    wb = openpyxl.load_workbook(filepath)

    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Apply header formatting to first row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Number formatting for numeric columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if abs(cell.value) < 1:
                        cell.number_format = '0.0000'
                    elif abs(cell.value) < 100:
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '#,##0.00'

    wb.save(filepath)


def export_metrics_to_csv(all_metrics: Dict[str, Dict[str, float]], output_dir: str):
    """Export metrics for each scenario to separate CSV files"""

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    for scenario_name, metrics in all_metrics.items():
        df = pd.DataFrame([metrics]).T
        df.columns = ['Value']
        df.index.name = 'Metric'

        csv_path = output_path / f'metrics_{scenario_name}.csv'
        df.to_csv(csv_path)
        print(f"Metrics CSV exported: {csv_path}")
